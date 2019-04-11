import numpy as np
import csv
import pickle
import os
import shutil
import gridfs
import openpyxl
from openpyxl import load_workbook
from copy import deepcopy

from Calculus.calculus import Calculus
from Calculus.CalcVar import CalcVar
import Config.variables as variables
import Config.config as config

from Scripts.dispatchable_indep import SequencePrediction
from Scripts.DemandPilotable import TotalDemandPilotable
from Scripts.HydroPilotable import HydroPilotable
from Data.DB.Mongo.mongo import Mongo
from Utils.Numpy import rollingMean, div0


if __debug__:
    import logging
    logger = logging.getLogger(__name__)
else:
    import warnings

VARS_FROMMONGO = {
    'ind_bz': ['Demand_Industry_Consumption_Efficiency_BZ_indeff_indtot_consumption'],
    'ind_ca': ['Demand_Industry_Consumption_Efficiency_CA_indeff_indtot_consumption'],
    'ind_cb': ['Demand_Industry_Consumption_Efficiency_CB_indeff_indtot_consumption'],
    'ind_cc': ['Demand_Industry_Consumption_Efficiency_CC_indeff_indtot_consumption'],
    'ind_cd': ['Demand_Industry_Consumption_Efficiency_CD_indeff_indtot_consumption'],
    'ind_ce': ['Demand_Industry_Consumption_Efficiency_CE_indeff_indtot_consumption'],
    'ind_cf': ['Demand_Industry_Consumption_Efficiency_CF_indeff_indtot_consumption'],
    'ind_cg': ['Demand_Industry_Consumption_Efficiency_CG_indeff_indtot_consumption'],
    'ind_ch': ['Demand_Industry_Consumption_Efficiency_CH_indeff_indtot_consumption'],
    'ind_ci': ['Demand_Industry_Consumption_Efficiency_CI_indeff_indtot_consumption'],
    'ind_cj': ['Demand_Industry_Consumption_Efficiency_CJ_indeff_indtot_consumption'],
    'ind_ck': ['Demand_Industry_Consumption_Efficiency_CK_indeff_indtot_consumption'],
    'ind_cl': ['Demand_Industry_Consumption_Efficiency_CL_indeff_indtot_consumption'],
    'ind_cm': ['Demand_Industry_Consumption_Efficiency_CM_indeff_indtot_consumption'],
    'res_aircond': ['Summary_Residential_Principal_Airconditioning_Consumption_Allvectors_EOD'],
    'res_airing': ['Summary_Residential_Principal_Airing_Consumption_Allvectors_EOD'],
    'res_cooking': ['Summary_Residential_Principal_Cooking_Consumption_Electricitygrid_EOD'],
    'res_ecs': ['Summary_Residential_Principal_Waterheating_Consumption_Electricitygrid_EOD'],
    'res_eq': ['Summary_Residential_Principal_Equipments_Consumption_Allvectors_EOD'],
    'res_heating': ['Summary_Residential_Principal_Heating_Consumption_Electricitygrid_EOD'],
    'res_lighting': ['Summary_Residential_Principal_Lighting_Consumption_Allvectors_EOD'],
    'ter_aircond': ['Summary_Tertiary_Airconditioning_Consumption_Allvectors_EOD'],
    'ter_eq': ['Summary_Tertiary_Specific_Consumption_Allvectors_EOD'],
    'ter_heating': ['Summary_Tertiary_Heating_Consumption_Electricitygrid_EOD'],
    'transport_cars': ['Demand_Transport_People_Consumption_electricitygrid_consumption_car'],
}


LIST_VAR_DEMAND = ['agr',
                   'ind_bz',
                   'ind_ca',
                   'ind_cb',
                   'ind_cc',
                   'ind_cd',
                   'ind_ce',
                   'ind_cf',
                   'ind_cg',
                   'ind_ch',
                   'ind_ci',
                   'ind_cj',
                   'ind_ck',
                   'ind_cl',
                   'ind_cm',
                   'res_aircond',
                   'res_airing',
                   'res_cooking',
                   'res_ecs',
                   'res_eq',
                   'res_heating',
                   'res_lighting',
                   'ter_aircond',
                   'ter_eq',
                   'ter_heating',
                   'ter_lighting',
                   'transport_cars',
                   'transport_other']

MAPPING_SECTORS = {'res': ['res_aircond',
                           'res_lighting',
                           'res_eq',
                           'res_heating',
                           'res_airing',
                           'res_cooking',
                           'res_ecs'],
                   'ter': ['ter_lighting',
                           'ter_heating',
                           'ter_eq',
                           'ter_aircond'],
                   'ind': ['ind_bz',
                           'ind_ca',
                           'ind_cb',
                           'ind_cc',
                           'ind_cd',
                           'ind_ce',
                           'ind_cf',
                           'ind_cg',
                           'ind_ch',
                           'ind_ci',
                           'ind_cj',
                           'ind_ck',
                           'ind_cl',
                           'ind_cm', ],
                   'agr': ['agr'],
                   'transport': ['transport_cars',
                                 'transport_other']}

VARS_TOSMOOTH = {'res_aircond': 72,
                 'ter_aircond': 72}
RENORM_WITHCONST = {'res_aircond': 0.75,
                    'ter_aircond': 0.5}
VARS_BASE_AND_THERM = ['res_aircond', 'ter_aircond']
SECTORS_OTHER = ['ter_', 'res_']
HOURLY_FROM_FILE = {'transport_other': 'datas/csv/for_eod/transport_other.csv'}


THERMOSENSIBLE_VARS = ['res_heating', 'res_aircond', 'ter_heating', 'ter_aircond']
listsupply = ['hydro_step',
              'Fioul',
              'import',
              'hydro_run',
              'Nucleaire',
              'Biomasse',
              'Charbon',
              'pv',
              'export',
              'hydro_lake',
              'UIOM',
              'Gaz',
              'wind_onshore',
              'wind_offshore']
list_pilot = ['pilotable_batterie',
              'pilotable_ind',
              'pilotable_res_ecs',
              'pilotable_res_eq',
              'pilotable_res_ter_heating',
              'pilotable_transport_cars', ]
testjson = {'parameters': [{myvar: 1 for myvar in LIST_VAR_DEMAND}]}
testjson['parameters'][0]['simulation_name'] = 'TEST'
testjson['simulation_id'] = 'abc'
testjson['parameters'][0]['weather_years'] = [2012]

testjson['parameters'][0].update({el: 1 for el in listsupply})
testjson['parameters'][0].update({el: 1 for el in list_pilot})


configparams = {'simulation_id': 'simu',
                'parameters': {'simulation_name': 4,
                               # Residential
                               'res_heating': 7,
                               'res_ecs': 8,
                               'res_lighting': 9,
                               'res_eq': 10,
                               'res_aircond': 11,
                               'res_airing': 12,
                               'res_cooking': 13,
                               # Tertiary
                               'ter_heating': 17,
                               'ter_lighting': 18,
                               'ter_eq': 19,
                               'ter_aircond': 20,
                               # Industry
                               'ind_bz': 24,
                               'ind_ca': 25,
                               'ind_cb': 26,
                               'ind_cc': 27,
                               'ind_cd': 28,
                               'ind_ce': 29,
                               'ind_cf': 30,
                               'ind_cg': 31,
                               'ind_ch': 32,
                               'ind_ci': 33,
                               'ind_cj': 34,
                               'ind_ck': 35,
                               'ind_cl': 36,
                               'ind_cm': 37,
                               # Transport
                               'transport_cars': 41,
                               'transport_other': 42,
                               # Agriculture
                               'agr': 46,
                               # Pilotable
                               'pilotable_batterie': 49,
                               'pilotable_ind': 50,
                               'pilotable_res_ecs': 51,
                               'pilotable_res_eq': 52,
                               'pilotable_res_ter_heating': 53,
                               'pilotable_transport_cars': 54,
                               # Production
                               'Nucleaire': 57,
                               'Gaz': 58,
                               'UIOM': 59,
                               'Biomasse': 60,
                               'hydro_run': 61,
                               'hydro_lake': 62,
                               'hydro_step': 63,
                               'wind_onshore': 64,
                               'wind_offshore': 65,
                               'pv': 66,
                               'import': 67,
                               'export': 68}}


class EOD(Calculus):
    '''
    Calcul de l'equilibre offre demande
    '''

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Always skip hash/cache because Simulation are meta-calculus
        self.folderpath = 'software/Dump/'
        self._skip_cache = True
        self._skip_hash = True
        self.hardcoded_years = [2015]
        self._mongo = Mongo(database=config.MONGO_SIMULATIONS_DB)
        self.to_save = {'Biomasse': 'Supply_Electricitygrid_Dispatchable_Biomass_All_prod_biomass',
                        'Charbon': 'Supply_Electricitygrid_Dispatchable_Coal_All_prod_coal',
                        'Fioul': 'Supply_Electricitygrid_Dispatchable_Oil_All_prod_oil',
                        'Gaz': 'Supply_Electricitygrid_Dispatchable_Gas_All_prod_gas',
                        'Nucleaire': 'Supply_Electricitygrid_Dispatchable_Nuclear_All_prod_nuc',
                        'UIOM': 'Supply_Electricitygrid_Dispatchable_Waste_All_prod_waste',
                        'hydro_run': 'Supply_Electricitygrid_Fatal_HydroRun_All_prod_hydrorun',
                        'hydrolake': 'Supply_Electricitygrid_Dispatchable_HydroLake_All_prod_hydrolake',
                        'hydrostep': 'Supply_Electricitygrid_Dispatchable_HydroStep_All_prod_hydrostep',
                        'intercoraw': 'Supply_Electricitygrid_Dispatchable_Rawinterco',
                        'tot_interco': 'Supply_Electricitygrid_Dispatchable_Interco_All_prod_interco',
                        'overconso': 'Supply_Electricitygrid_Dispatchable_Overconso_All_prod_overconso',
                        'pilotable_batterie': 'Supply_Electricitygrid_Pilot_Batteries',
                        'pilotable_ind': 'Supply_Electricitygrid_Pilot_Industry',
                        'pilotable_res_ecs': 'Supply_Electricitygrid_Pilot_Residential_Waterheating',
                        'pilotable_res_eq': 'Supply_Electricitygrid_Pilot_Residential_Equipments',
                        'pilotable_res_ter_heating': 'Supply_Electricitygrid_Pilot_Residential_Tertiary_Heating',
                        'pilotable_transport_cars': 'Supply_Electricitygrid_Pilot_Electric_Cars',
                        'pv': 'Supply_Electricitygrid_Fatal_PV_All_prod_pv',
                        'underconso': 'Supply_Electricitygrid_Dispatchable_Overprod_All_prod_overprod',
                        'wind_offshore': 'Supply_Electricitygrid_Fatal_Wind_offshore_All_prod_wind',
                        'wind_onshore': 'Supply_Electricitygrid_Fatal_Wind_All_prod_wind',

                        'res_aircond': 'Summary_Residential_Airconditioning_Consumption_Allvectors_EOD',
                        'res_airing': 'Summary_Residential_Airing_Consumption_Allvectors_EOD',
                        'res_cooking': 'Summary_Residential_Cooking_Consumption_Electricitygrid_EOD',
                        'res_ecs': 'Summary_Residential_Waterheating_Consumption_Electricitygrid_EOD',
                        'res_eq': 'Summary_Residential_Equipments_Consumption_Allvectors_EOD',
                        'res_heating': 'Summary_Residential_Heating_Consumption_Electricitygrid_EOD',
                        'res_lighting': 'Summary_Residential_Lighting_Consumption_Allvectors_EOD',
                        'ter_aircond': 'Summary_Tertiary_Airconditioning_Consumption_Allvectors_EOD',
                        'ter_eq': 'Summary_Tertiary_Specific_Consumption_Allvectors_EOD',
                        'ter_heating': 'Summary_Tertiary_Heating_Consumption_Electricitygrid_EOD',
                        'transport_cars': 'Demand_Transport_People_Consumption_electricitygrid_consumption_car',
                        'agr': 'Summary_Agriculture_Electricitygrid_Consumption_EOD',
                        'ind_tot': 'Summary_Industry_Consumption_Electricitygrid_EOD',
                        'transport_tot': 'Summary_Transport_Consumption_Electricitygrid_EOD',
                        'total_conso': 'Summary_Electricitygrid_Consumption_EOD',
                        'demand_to_match': 'Summary_Electricitygrid_Consumption_to_Dispatchable_EOD',

                        'internal_conso': 'Summary_Demand_Internal_Consumption',
                        'losses': 'Electric_Loss_EOD',
                        'total_demand': 'Summary_Electricitygrid_Consumption_Losses_Internal',
                        'total_lisse': 'Summary_Electricitygrid_Consumption_Net_Pilot', }

    def build_from_xls(self, filename):
        toupdate = {'Fioul': 0.0001, 'Charbon': 0.0001, }
        wb = load_workbook(filename='Template EOD.xlsx', read_only=True)
        ws = wb['Sheet1']
        nsimu = ws.max_column - 2
        storeconfig = deepcopy(configparams)
        storeconfig['parameters'] = [{} for i in range(nsimu)]
        listrows = list(ws.rows)
        for key, idx in configparams['parameters'].items():
            row = listrows[idx - 1]
            for i, subl in enumerate(storeconfig['parameters']):
                subl[key] = row[i + 2].value
        for subl in storeconfig['parameters']:
            subl.update(toupdate)
        return storeconfig

    def _run(self, filename='Template EOD.xlsx'):
        request = self.build_from_xls(filename)
        parameters = request['parameters']
        simuid = request['simulation_id']
        self.storage = {}
        for onesimu in parameters:
            annual_vals_demand, scenars, annual_vals_supply, pilot_param = self.build_annual_vals(onesimu)
            scenar = list(scenars.keys())[0]
            self.storage[scenar] = {}
            for year in self.hardcoded_years:
                self.storage[scenar][int(year)] = {}
                coll_name = simuid + '_' + onesimu['simulation_name'] + '_' + str(year)
                coll_name = coll_name.replace(' ', '_')
                demand_builder = HourlyDemandBuilder(int(year), annual_vals_demand, scenars)
                demand_builder.load_datas()
                netbuilder = self.calculus('Calculus.EOD.eod.NetConsoBuilder',
                                           int(year), demand_builder, annual_vals_supply, pilot_param)
                self.net_builder = netbuilder
                netbuilder.build_data()
                self.net_builder = netbuilder
                self.save_vars(netbuilder, coll_name)
                for myvar in netbuilder.net_scenars[scenar]:
                    self.storage[scenar][int(year)][myvar] = netbuilder.net_scenars[scenar][myvar]
        try:
            grapher = self.calculus('Calculus.GraphCsvZip.FolderBuilder.FolderADEME',
                                    simuid, rootpath=self.folderpath)
        except RuntimeWarning as e:
            logger.warn("Runtime warning : %s", e)
            import warnings
            warnings.filterwarnings('ignore')
            grapher = self.calculus('Calculus.GraphCsvZip.FolderBuilder.FolderADEME',
                                    simuid, rootpath=self.folderpath)
        grapher(self.storage)
        rootpath = grapher.rootpath
        folderzip = self.folderpath
        if not os.path.exists(folderzip):
            os.makedirs(folderzip)
        with open(rootpath + 'summary_annual_hypo.csv', 'w') as f:
            writer = csv.writer(f)
            for scenar1 in parameters:
                header = ['varname' + ' - ' + scenar1['simulation_name'], '2015']  # '2012', '2013', '2015']
                writer.writerow(header)
                listkeys1 = sorted(list(scenar1))
                for key in listkeys1:
                    if key in self.storage[scenar1['simulation_name']][self.hardcoded_years[0]]:
                        row = [key] + [self.storage[scenar1['simulation_name']][year]
                                       [key].sum() / 1e12 for year in self.hardcoded_years]
                        writer.writerow(row)
        shutil.make_archive(folderzip + simuid, 'zip', rootpath)
        # cleanup temp file
        shutil.rmtree(rootpath)
        with open(folderzip + simuid + '.zip', 'rb') as f:
            data = f.read()
            self._mongo.set_db_coll(database=config.MONGO_ZIP_DB, collection=config.__ZIP_GRAPHS__)
            fs = gridfs.GridFS(self._mongo._database)
            stored = fs.put(data)
            doc = {'fs_id': stored, 'name': simuid}
            self._mongo._insert(doc, database=config.MONGO_ZIP_DB, collection=config.__ZIP_GRAPHS__)
        return None

    def save_vars(self, netbuilder, coll_name):
        listdocs = []
        for key, varname in self.to_save.items():
            scenar = list(netbuilder.scenars.keys())[0]
            map2 = {'projection': 'map_2', 'varname': varname, 'year': int(coll_name.split('_')[-1])}
            map2['FR99999'] = self.map_builder_from_array(netbuilder.net_scenars[scenar][key])
            curve2 = {'projection': 'curve_2', 'varname': varname, 'year': int(coll_name.split('_')[-1])}
            curve2['FR99999'] = list(netbuilder.net_scenars[scenar][key])
            listdocs.append(map2)
            listdocs.append(curve2)
        self._mongo.insert(documents=listdocs, flag_raw_input=1,
                           collection=coll_name, database=config.MONGO_SIMULATIONS_DB)

    def map_builder_from_array(self, myarray):
        months = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        ret = [myarray.sum()]
        idx_i = 0
        for month in months:
            idx_f = idx_i + month * 24
            ret.append(myarray[idx_i:idx_f].sum())
            idx_i = idx_f
        for i in range(52):
            ret.append(myarray[i * 7 * 24:(i + 1) * 7 * 24].sum())
        return ret

    def build_annual_vals(self, onesimu):
        annual_vals_demand = onesimu.copy()
        annual_vals_supply = onesimu.copy()
        tempkeys = []
        for key in annual_vals_demand:
            if key not in LIST_VAR_DEMAND:
                tempkeys.append(key)
        for key in tempkeys:
            del annual_vals_demand[key]
        pilot_param = {key: max(onesimu[key], 0.0001) for key in list_pilot}
        scenars = {onesimu['simulation_name']: None}
        return annual_vals_demand, scenars, {key: onesimu[key] for key in listsupply}, pilot_param


class HourlyDemandBuilder(Calculus):

    def __init__(self, bilan_year, annual_vals, scenars, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.bilan_year = bilan_year
        self.bilan_years = [2015]
        self._mongo = Mongo(database=config.MONGO_SIMULATIONS_DB, collection='bilan' + str(bilan_year))
        self.scenars = scenars
        self.annual_vals = {scenar: annual_vals for scenar in self.scenars}
        self.aircondargs = {'ener_ventil': 0.3,
                            'ener_pure_thermo': 0.1,
                            'tempcutoff': 15,
                            'tempsmooth': 72,
                            'pure_smooth': 10,
                            'tempsensitivity': 0.5}
        self.aircondargs['basalaircond_frac'] = 1 - \
            self.aircondargs['ener_ventil'] - self.aircondargs['ener_pure_thermo']

    def load_datas(self):
        self.load_hourly_vals()
        self.thermosensible_ref()
        self.build_renormalized_data()

    def thermosensible_ref(self):
        self.temperatures = {key: None for key in self.bilan_years}
        self.smoothed_temp = {key: None for key in self.bilan_years}
        self.ref_thermosensible_yearly = {myvar: [] for myvar in THERMOSENSIBLE_VARS}
        for myvar in THERMOSENSIBLE_VARS:
            for oneyear in self.bilan_years:
                query = {"varname": {'$in': [varname for varname in VARS_FROMMONGO[myvar]]},
                         "projection": 'map_2'}
                docs = self._mongo.find(collection='bilan' + str(oneyear), query=query)
                temp = 0
                for doc in docs:
                    temp += doc['FR99999'][0]
                self.ref_thermosensible_yearly[myvar].append(temp)
                query = {"varname": 'Global_Weather_Temperature', "projection": 'curve_2'}
                docs = self._mongo.find(collection='bilan' + str(oneyear), query=query)
                self.temperatures[oneyear] = np.array(docs[0]['FR99999'])
                self.smoothed_temp[oneyear] = rollingMean(self.temperatures[oneyear], self.aircondargs['tempsmooth'])
        self.mean_thermosensible = {key: sum(val) / len(self.bilan_years) for key, val in self.ref_thermosensible_yearly.items()}
        self._mongo.set_db_coll(collection='bilan' + str(self.bilan_year))

    def load_hourly_vals(self):
        one_scenar = list(self.scenars.keys())[0]
        self.ref_hourly = {myvar: None for myvar in self.annual_vals[one_scenar]}
        for myvar in self.annual_vals[one_scenar]:
            if myvar in VARS_FROMMONGO:
                query = {"varname": {'$in': [varname for varname in VARS_FROMMONGO[myvar]]},
                         "projection": 'curve_2'}
                docs = self._mongo.find(query=query)
                temp_hourly = np.zeros((8760,))
                for doc in docs:
                    temp_hourly += np.array(doc['FR99999'])
            elif myvar in HOURLY_FROM_FILE:
                with open(HOURLY_FROM_FILE[myvar], 'r') as f:
                    reader = csv.reader(f)
                    for r_idx, line in enumerate(reader):
                        if r_idx == 0:
                            temp_hourly = np.array([float(el) for el in line])
            else:
                temp_hourly = np.ones((8760,))
            self.ref_hourly[myvar] = temp_hourly

    def moving_average(self, a, n=72):
        if n % 2 != 0:
            raise
        ret = np.cumsum(a, dtype=float)
        ret[n:] = ret[n:] - ret[:-n]
        ret = np.concatenate((a[0:int(n / 2)], ret[n - 1:] / n, a[- int(n / 2) + 1:]), axis=0)
        return ret / ret.sum() * a.sum()

    def build_renormalized_data(self):
        self.renormalized_data = {}
        for scenar in self.scenars.keys():
            self.renormalized_data[scenar] = {}
            self.renormalized_data[scenar]['total_conso'] = np.zeros((8760,))
            for myvar in self.annual_vals[scenar]:
                if myvar in THERMOSENSIBLE_VARS:
                    if self.annual_vals[scenar][myvar]:
                        self.renormalized_data[scenar][myvar] = div0(self.ref_hourly[myvar], self.mean_thermosensible[
                                                                     myvar]) * self.annual_vals[scenar][myvar] * 1e12
                    else:
                        self.renormalized_data[scenar][myvar] = self.ref_hourly[myvar]
                else:
                    if self.annual_vals[scenar][myvar]:
                        self.renormalized_data[scenar][myvar] = div0(self.ref_hourly[myvar], self.ref_hourly[
                                                                     myvar].sum()) * self.annual_vals[scenar][myvar] * 1e12
                    else:
                        self.renormalized_data[scenar][myvar] = self.ref_hourly[myvar]
                if myvar in VARS_BASE_AND_THERM:
                    if self.annual_vals[scenar][myvar]:
                        tempval = self.annual_vals[scenar][myvar]
                    else:
                        tempval = self.renormalized_data[scenar][myvar].sum() / 1e12
                    basal = np.ones((8760,)) / 8760 * \
                        tempval * 1e12 * self.aircondargs['ener_ventil']
                    basalaircond = np.maximum(
                        self.smoothed_temp[self.bilan_year] - self.aircondargs['tempcutoff'], 0) * self.aircondargs['tempsensitivity']
                    basalaircond = basalaircond / basalaircond.sum()
                    basalaircond = basalaircond * \
                        self.aircondargs['basalaircond_frac'] * tempval * 1e12
                    aircond = rollingMean(self.renormalized_data[scenar][myvar], self.aircondargs[
                                          'pure_smooth']) * self.aircondargs['ener_pure_thermo']
                    self.renormalized_data[scenar][myvar] = basal + basalaircond + aircond
            # temp_notothers = {key: np.zeros((8760,)) for key in SECTORS_OTHER}
            # temp_others = {}
            # for myvar in self.annual_vals[scenar]:
            #     if SECTORS_OTHER[0] in myvar:
            #         if 'other' in myvar:
            #             temp_others[myvar] = self.renormalized_data[scenar][myvar]
            #         elif 'heating' not in myvar and 'aircond' not in myvar:
            #             temp_notothers[SECTORS_OTHER[0]] += self.renormalized_data[scenar][myvar]
            #         elif 'heating' in myvar:
            #             temp_notothers[SECTORS_OTHER[0]] += 0.5 * self.renormalized_data[scenar][myvar]
            #     elif SECTORS_OTHER[1] in myvar:
            #         if 'other' in myvar:
            #             temp_others[myvar] = self.renormalized_data[scenar][myvar]
            #         elif 'heating' not in myvar and 'aircond' not in myvar:
            #             temp_notothers[SECTORS_OTHER[1]] += self.renormalized_data[scenar][myvar]
            #         elif 'heating' not in myvar:
            #             temp_notothers[SECTORS_OTHER[1]] += 0.5 * self.renormalized_data[scenar][myvar]
            # for myvar in temp_others:
            #     otherkey = myvar.split('_')[0] + '_'
            #     self.renormalized_data[scenar][myvar] = temp_notothers[otherkey] / \
            #         temp_notothers[otherkey].sum() * temp_others[myvar].sum()
            for myvar in self.annual_vals[scenar]:
                self.renormalized_data[scenar]['total_conso'] += self.renormalized_data[scenar][myvar]
            self.renormalized_data[scenar]['losses'] = self.renormalized_data[scenar][
                'total_conso'] * (0.06 + (self.renormalized_data[scenar]['total_conso'] / 1e9 / 100) * 0.04)
            self.renormalized_data[scenar]['internal_conso'] = np.ones((8760,)) / 8760 * (15 * 1e12 / 8760)
            tosum = ['total_conso', 'losses', 'internal_conso']
            self.renormalized_data[scenar]['total_demand'] = sum(self.renormalized_data[scenar][el] for el in tosum)
            self.renormalized_data[scenar]['ind_tot'] = sum(
                self.renormalized_data[scenar][el] for el in MAPPING_SECTORS['ind'])
            self.renormalized_data[scenar]['transport_tot'] = self.renormalized_data[
                scenar]['transport_cars'] + self.renormalized_data[scenar]['transport_other']


class MarineWind(Calculus):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geo_idx = self._cache.get_val('geocodes_indexes')
        self.wind_geo = ["FR44184",
                         "FR22278",
                         "FR14191",
                         "FR76259",
                         "FR76711",
                         "FR85163"]
        self.wind_idx = [self.geo_idx[geo] for geo in self.wind_geo]
        self.vestasV112_3450 = [3.5, 11, 25]

    def build_charge_marine(self, year, std_div=2):
        weather_inst = self.calculus('Calculus.Weather.Hypotheses.WeatherFaithful')
        myweather = self.get_variable(weather_inst([year], [year]))
        queried_geo = myweather['interpolation'][year]['AverageWind'][self.wind_idx, :]
        self.meanwind = np.array(np.dot(queried_geo, myweather['data'][year]['AverageWind'])).mean(axis=0)
        mean_ini = self.meanwind.mean()
        self.treated_wind = (self.meanwind - mean_ini) / std_div + mean_ini

    def compute_wind_shift(self, w_coeff=1.53):
        coeff = 1 / (self.vestasV112_3450[1] - self.vestasV112_3450[0])
        cutoff_array = 0.5 * (np.sign(self.vestasV112_3450[2] - self.treated_wind * w_coeff) + 1)
        percent_prod = np.minimum(np.maximum(0, coeff * (self.treated_wind * w_coeff -
                                                         self.vestasV112_3450[0])), 1) * cutoff_array
        self.offshorewind = percent_prod
        return percent_prod.mean()

ENRVARS = {'wind_onshore': ['Supply_Electricitygrid_Fatal_Wind_All_prod_wind', 10324],
           'pv': ['Supply_Electricitygrid_Fatal_PV_All_prod_pv', 6196],
           'hydro_run': ['Supply_Electricitygrid_Fatal_HydroRun_All_prod_hydrorun', 9985]}

ECHPHY = 'Supply_Electricitygrid_Dispatchable_Interco_All_prod_interco'

FATAL_TO_REMOVE = ['pv',
                   'wind_onshore',
                   'wind_offshore',
                   'hydro_run', ]
charge_coef = [1.52, 1.18, 1.0, 1.0, 1.0, 1.0]
charge_target = [14.7, 24.3, 42, 1, 1, 1]
charge_ref = [9.7, 20.7, 42.1, 1.0, 1.0, 1.0]


class NetConsoBuilder(Calculus):
    '''
    - check self.scenars idx value not used

    - integrer la creation des indicatrices defaut et interco a la hausse et a la baisse

    '''

    def __init__(self, bilan_year, demand_builder, annual_vals, pilot_param, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.annual_vals = annual_vals
        self.nucmin = 0.3
        self._mongo = Mongo(database=config.MONGO_SIMULATIONS_DB, collection='bilan' + str(bilan_year))
        self.rawpickle_data = 'datas/csv/for_eod/dumptot_supplydata'
        self.scenars = demand_builder.scenars
        self.fataldata = {scenar: {} for scenar in self.scenars}
        self.demand_builder = demand_builder
        self.wind_offshore_builder = self.calculus('Calculus.EOD.eod.MarineWind')
        self.wind_offshore_builder.build_charge_marine(bilan_year)
        _ = self.wind_offshore_builder.compute_wind_shift()
        self.pmax_ref = {'Fioul': 6942.0,
                         'Charbon': 6395.0,
                         'Gaz': 8040.0,
                         'Nucleaire': 63130.0,
                         'Biomasse': 913.8,
                         'UIOM': 416.5,
                         'Solaire': 6196,
                         'Éolien': 10324}
        self.kwargs_pilot = {'val_min_pic': {key: -val * 1e9 for key, val in pilot_param.items()},
                             'val_max_pic': {key: val * 1e9 for key, val in pilot_param.items()}, }
        self.smooth_coef = 1

        # INTERCO!!!

    def build_net_conso(self):
        self.net_scenars = {scenar: {} for scenar in self.scenars}
        for scenar in self.scenars:
            net = self.demand_builder.renormalized_data[scenar]['total_demand']
            for fatal, coef in zip(FATAL_TO_REMOVE, charge_coef):
                self.fataldata[scenar][fatal] *= coef
                net = net - self.fataldata[scenar][fatal]
        # attention au signe
            net = net + self.raw_ech[scenar]
            self.net_scenars[scenar]['net_demand'] = net
            for myvar, myval in self.demand_builder.renormalized_data[scenar].items():
                self.net_scenars[scenar][myvar] = myval
            for myvar in FATAL_TO_REMOVE:
                self.net_scenars[scenar][myvar] = self.fataldata[scenar][myvar]
            self.net_scenars[scenar]['intercoraw'] = self.raw_ech[scenar]

    def build_net_smoothed(self):
        pilotable = TotalDemandPilotable(**self.kwargs_pilot)
        for scenar in self.net_scenars:
            restemp = pilotable.GetLissage(self.net_scenars[scenar])
            res = {'total_lisse': restemp['total_lisse']}
            for key in restemp['deltas']:
                res[key] = restemp['deltas'][key]
            self.net_scenars[scenar].update(res)
            hpmax = {'hydrostep': self.pmax[scenar]['hydro_step'],
                     'hydrolake': self.pmax[scenar]['hydro_lake']}
            hydro_dyn = HydroPilotable(newpmax=hpmax)
            res2 = hydro_dyn.GetLissage(self.net_scenars[scenar]['total_lisse'])
            self.net_scenars[scenar].update(res2)
            self.net_scenars[scenar]['demand_post_hydro'] = self.net_scenars[
                scenar]['total_lisse'] - sum(onehydro for onehydro in res2.values())

    def compute_overshoot_envelope(self):
        for scenar in self.net_scenars:
            self.net_scenars[scenar]['max_envelope_disp'] = (self.kmax[scenar]['total']) * 0.99
            self.net_scenars[scenar]['min_envelope_disp'] = (
                self.kmin[scenar]['total']) + 0.01 * self.net_scenars[scenar]['max_envelope_disp']
            # to W
            self.net_scenars[scenar]['max_envelope_disp'] = self.net_scenars[scenar]['max_envelope_disp'] * 1e6
            self.net_scenars[scenar]['min_envelope_disp'] = self.net_scenars[scenar]['min_envelope_disp'] * 1e6
            # over and under conso
            self.net_scenars[scenar]['overconso'] = np.maximum(
                self.net_scenars[scenar]['demand_post_hydro'] - self.net_scenars[scenar]['max_envelope_disp'], 0)
            self.net_scenars[scenar]['underconso'] = np.minimum(
                self.net_scenars[scenar]['demand_post_hydro'] - self.net_scenars[scenar]['min_envelope_disp'], 0)
            # adjust interco to absorb as much of these as possible
            raw_adjusted = self.raw_ech[scenar] - \
                self.net_scenars[scenar]['underconso'] - self.net_scenars[scenar]['overconso']
            adjusted = np.maximum(np.minimum(self.pmax[scenar]['export']
                                             * 1e6, raw_adjusted), -self.pmax[scenar]['import'] * 1e6)
            self.net_scenars[scenar]['deltainterco'] = adjusted - self.raw_ech[scenar]
            self.net_scenars[scenar]['default_sure'] = np.round(self.net_scenars[scenar]['overconso'] +
                                                                self.net_scenars[scenar]['underconso'] + self.net_scenars[scenar]['deltainterco'])
            self.net_scenars[scenar]['demand_to_match'] = self.net_scenars[scenar]['demand_post_hydro'] - \
                self.net_scenars[scenar]['overconso'] - self.net_scenars[scenar]['underconso']

    def compute_eod(self):
        for scenar in self.net_scenars:
            myseq = SequencePrediction(6)
            margs = self.prepare_vars_for_optim(scenar)
            mydispatchable = myseq.predict('pred', *margs)
            for idx, key in enumerate(self.order):
                self.net_scenars[scenar][key] = mydispatchable[:, idx] * 1e9
                self.net_scenars[scenar]['dispodispatch'] = {'K_max': margs[0], 'K_min': margs[1]}
                # self.net_scenars[scenar][key] = margs[2] * \
                #     np.expand_dims(margs[0][:, idx] / margs[0].sum(axis=1), axis=1) * 1e9
            self.net_scenars[scenar]['disp_pmax'] = {key: val for key, val in self.pmax[scenar].items()}

    def prepare_vars_for_optim(self, scenar):
        scale_factor = 10**9
        self.order = ['Fioul', 'Charbon', 'Gaz', 'Nucleaire', 'Biomasse', 'UIOM']
        K_max = np.concatenate(tuple(np.expand_dims(self.kmax[scenar][
                               key], axis=1) for key in self.order), axis=1) / 1e3
        K_min = np.concatenate(tuple(np.expand_dims(self.kmin[scenar][
                               key], axis=1) for key in self.order), axis=1) / 1e3
        asked = np.expand_dims(self.net_scenars[scenar]['demand_to_match'] / scale_factor, axis=1)
        sumpmax = sum(self.pmax[scenar][key] for key in self.order)
        seq = np.concatenate(tuple(
            asked * self.pmax[scenar][key] / sumpmax for key in self.order), axis=1)
        order_old = ['Fioul', 'Charbon', 'Gaz', 'Nucleaire', 'Biomasse', 'UIOM']
        static_costs1 = np.array([83.773, 151.362, 50.419, 17.166, 30.439, 30.141]) / scale_factor
        stat_c = static_costs1[[order_old.index(key) for key in self.order]]
        dynamic_costs1 = np.array([0.145, 0.371, 0.086, 0.312, 3.77, 5.295])
        dyn_c = dynamic_costs1[[order_old.index(key) for key in self.order]]
        nhour_start = np.array([3.5, 3.5, 2, 10, 3.5, 3.5])
        nh_start = nhour_start[[order_old.index(key) for key in self.order]]
        start_c = np.array([static_costs1[i] * nhour_start[i] * K_max[:, i].mean() for i in range(6)])
        poff_c = - 0.0 * start_c
        return K_max, K_min, asked, seq, stat_c, dyn_c, start_c, poff_c, np.zeros((6,))

    def build_data(self):
        self.load_capa()
        self.build_from_bilan()
        self.offshore_prod()
        self.build_net_conso()
        self.build_kdisp()
        self.build_net_smoothed()
        self.compute_overshoot_envelope()
        self.compute_eod()
        self.change_shape()
        self.add_interco()
        self.add_nuc_exp()
        self.add_kdkpku()

    def add_kdkpku(self):
        dictvars = {'Nucleaire': 3}
        for scenar in self.net_scenars:
            for myvar, idx in dictvars.items():
                self.net_scenars[scenar]['Kd'] = self.net_scenars[scenar]['dispodispatch'][
                    'K_max'][:, idx] / self.net_scenars[scenar]['disp_pmax'][myvar] * 1e3
                self.net_scenars[scenar]['Ku'] = self.net_scenars[scenar][myvar] / \
                    self.net_scenars[scenar]['dispodispatch']['K_max'][:, idx]
                self.net_scenars[scenar]['Kp'] = self.net_scenars[scenar][
                    myvar] / self.net_scenars[scenar]['disp_pmax'][myvar] * 1e3

    def add_nuc_exp(self):
        dictvars = {'Nucleaire': 3}
        for scenar in self.net_scenars:
            temp = self.net_scenars[scenar]['dispodispatch']['K_max'][:, dictvars['Nucleaire']] * 1e9
            disp_nuc = np.maximum(temp - self.net_scenars[scenar]['Nucleaire'], 0)
            dispinterco = self.net_scenars[scenar]['disp_pmax']['export'] * 1e6 - self.net_scenars[scenar]['intercoraw'] - \
                self.net_scenars[scenar]['deltainterco']
            self.net_scenars[scenar]['nuc_exp'] = np.minimum(disp_nuc, dispinterco)

    def add_interco(self):
        for scenar in self.net_scenars:
            self.net_scenars[scenar]['tot_interco'] = self.net_scenars[scenar][
                'intercoraw'] + self.net_scenars[scenar]['deltainterco']

    def change_shape(self):
        for scenar in self.net_scenars:
            for myvar in self.net_scenars[scenar]:
                if type(self.net_scenars[scenar][myvar]) == np.ndarray:
                    if self.net_scenars[scenar][myvar].shape == (8760, 1):
                        self.net_scenars[scenar][myvar] = np.ndarray.flatten(self.net_scenars[scenar][myvar])

    def offshore_prod(self):
        rescaled_data = {scenar: {} for scenar in self.scenars}
        for scenar in self.scenars:
            rescaled_data[scenar]['wind_offshore'] = self.wind_offshore_builder.offshorewind * \
                self.pmax[scenar]['wind_offshore'] * 1e6
            self.fataldata[scenar].update(rescaled_data[scenar])

    def build_from_bilan(self):
        rescaled_data = {scenar: {} for scenar in self.scenars}
        self.raw_ech = {scenar: None for scenar in self.scenars}
        for scenar in self.scenars:
            for key in ENRVARS:
                query = {"varname": ENRVARS[key][0], "projection": 'curve_2'}
                docs = self._mongo.find(query=query)
                rescaled_data[scenar][key] = np.array(docs[0]['FR99999']) / (ENRVARS[key][1]) * self.pmax[scenar][key]
            self.fataldata[scenar].update(rescaled_data[scenar])
            query = {"varname": ECHPHY, "projection": 'curve_2'}
            docs = self._mongo.find(query=query)
            self.raw_ech[scenar] = self.echphy_rescale(np.array(docs[0]['FR99999']), scenar)

    def _hydro(self):
        hydro_step = self.datadict['data'][:, 7]
        hydro_lake = self.datadict['data'][:, 6]
        hydro_step[:1000] = hydro_step[1000:2000]

    def build_kdisp(self):
        self.raw_data = {'disp': {}, 'asked': {}}
        self.dispo = {'K_max': {}, 'K_min': {}}
        with open(self.rawpickle_data, 'rb') as reader:
            self.datadict = pickle.load(reader)
        self.datadict['techno_order'][self.datadict['techno_order'].index('Nucléaire')] = 'Nucleaire'
        for key in self.datadict:
            if key not in ['techno_order']:
                self.datadict[key] = np.array(self.datadict[key])
        self.kmin = {scenar: {} for scenar in self.scenars}
        self.kmax = {scenar: {} for scenar in self.scenars}
        for scenar in self.scenars:
            for idx, key in enumerate(self.datadict['techno_order'][:6]):
                self.kmin[scenar][key] = self.datadict['K_min'][:, idx] * \
                    self.pmax[scenar][key] / self.datadict['K_max'][:, idx].max()
                self.kmax[scenar][key] = self.datadict['K_max'][:, idx] * \
                    self.pmax[scenar][key] / self.datadict['K_max'][:, idx].max()
            self.kmin[scenar]['Nucleaire'] = np.minimum(self.nucmin * self.kmax[scenar]['Nucleaire'], 10000)
            self.kmin[scenar]['Gaz'] = 0 * self.kmax[scenar]['Gaz']
            self.kmin[scenar]['total'] = sum(self.kmin[scenar][key] for key in self.datadict['techno_order'][:6])
            self.kmax[scenar]['total'] = sum(self.kmax[scenar][key] for key in self.datadict['techno_order'][:6])

    def echphy_rescale(self, echphy, scenar):
        return echphy * self.pmax[scenar]['export'] / (self.pmax[scenar]['export'] - 8000)

    def load_capa(self):
        self.pmax = {scenar: {} for scenar in self.demand_builder.scenars.keys()}
        scenar = list(self.pmax.keys())[0]
        for key, val in self.annual_vals.items():
            self.pmax[scenar][key] = max(val * 1e3, 0.1)
